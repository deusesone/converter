from openpyxl import load_workbook

data_file_path = 'inventory.xlsx'
file = load_workbook(data_file_path)
page = file['proxmox']
row = list(page.rows)
column = list(page.columns)

output_file = open ('inventory.yaml', 'w')

for i in range (1, page.max_row + 1):
    if i == 1:
        continue
    hostname = page.cell(row=i, column=1).value
    print(hostname)
    ip = page.cell(row=i, column=2).value
    print(ip)
    vm_id = page.cell(row=i, column=3).value
    print(vm_id)
    bridge = page.cell(row=i, column=4).value
    print(bridge)
    short_net = page.cell(row=i, column=5).value
    print(short_net)
    vm_storage = page.cell(row=i, column=6).value
    print(vm_storage)
    vm = page.cell(row=i, column=7).value
    print(vm)
    template_id = page.cell(row=i, column=8).value
    print(template_id)
    cores = page.cell(row=i, column=9).value
    print(cores)
    memory = page.cell(row=i, column=10).value
    print(memory)
    gateway = page.cell(row=i, column=11).value
    print(gateway)
    proxmox_host = page.cell(row=i, column=12).value
    print(proxmox_host)
    proxmox_ip = page.cell(row=i, column=13).value
    print(proxmox_ip)
    print(hostname + ' ansible_host=' + ip + ' proxmox_vm_id=' + str(vm_id) + ' proxmox_vm_net_bridge=' + bridge + ' ansible_short_net=' + str(short_net) + ' proxmox_vm_storage=' + vm_storage + ' vm=' + vm + ' proxmox_template_id=' + str(template_id) + ' proxmox_vm_cores=' + str(cores) + ' proxmox_vm_mem=' + str(memory) + ' ct_gw=' + str(gateway) + ' proxmox_vm_host=' + proxmox_host + ' proxmox_delegate_host=' + proxmox_ip, file=output_file)